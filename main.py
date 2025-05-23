from flask import Flask, request, jsonify
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
import random

app = Flask(__name__)
app.secret_key = 'your_secret_key'

EXCEL_PATH = 'bot.xlsx'
LOG_XLSX_PATH = 'chat_log.xlsx'

# 사용자 정보 조회
def get_user(id_code):
    df = pd.read_excel(EXCEL_PATH, sheet_name='인증', dtype=str).fillna('')
    user = df[df['Id_code'] == id_code]
    return user.iloc[0].to_dict() if not user.empty else None

# 로그 기록
def log_action(id_code, user_input, bot_response):
    user = get_user(id_code)
    name = user['Name'] if user and 'Name' in user else 'Unknown'
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    log_data = {
        'timestamp': [timestamp],
        'Id_code': [id_code],
        'Name': [name],
        'input': [user_input],
        'bot_response': [bot_response]
    }
    df_log = pd.DataFrame(log_data)

    if os.path.exists(LOG_XLSX_PATH):
        wb = load_workbook(LOG_XLSX_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if '전체로그' not in wb.sheetnames:
        ws_all = wb.create_sheet('전체로그')
        for r in dataframe_to_rows(df_log, index=False, header=True):
            ws_all.append(r)
    else:
        ws_all = wb['전체로그']
        for r in dataframe_to_rows(df_log, index=False, header=False):
            ws_all.append(r)

    safe_name = ''.join(c for c in name if c.isalnum() or c in ('_', '-')).strip() or 'Unknown'
    if safe_name not in wb.sheetnames:
        ws_user = wb.create_sheet(safe_name)
        for r in dataframe_to_rows(df_log, index=False, header=True):
            ws_user.append(r)
    else:
        ws_user = wb[safe_name]
        for r in dataframe_to_rows(df_log, index=False, header=False):
            ws_user.append(r)

    wb.save(LOG_XLSX_PATH)

# 트리형 조사
def investigate_tree(select_path, user_input):
    df = pd.read_excel(EXCEL_PATH, sheet_name='조사', dtype=str).fillna('')
    path = [x for x in select_path.split(',') if x] if select_path else []

    if user_input == "처음으로":
        path = []
    elif user_input == "이전으로" and path:
        path = path[:-1]
    elif user_input and user_input not in ["이전으로", "처음으로"]:
        path.append(user_input)

    depth = len(path)
    cond = True
    for i, sel in enumerate(path):
        cond = cond & (df[f'장소{i+1}'] == sel)
    next_col = f'장소{depth+1}' if depth < 5 else '타겟'
    next_options = sorted(df[cond][next_col].dropna().unique())

    if depth == 6 or (depth == 5 and user_input and user_input not in ["이전으로", "처음으로"]):
        cond2 = cond & (df[next_col] == (user_input if user_input not in ["이전으로", "처음으로"] else path[-1]))
        row = df[cond2]
        if not row.empty:
            msg = row.iloc[0]['출력']
            quick_replies = [
                {"label": "이전으로", "action": "message", "messageText": "이전으로"},
                {"label": "처음으로", "action": "message", "messageText": "처음으로"}
            ]
        else:
            msg = "해당 조건에 맞는 장소가 없습니다."
            quick_replies = [
                {"label": "이전으로", "action": "message", "messageText": "이전으로"},
                {"label": "처음으로", "action": "message", "messageText": "처음으로"}
            ]
    else:
        msg = f"조사가능: {', '.join(next_options)}"
        quick_replies = [
            {"label": opt, "action": "message", "messageText": opt} for opt in next_options
        ]
        if path:
            quick_replies.append({"label": "이전으로", "action": "message", "messageText": "이전으로"})
        quick_replies.append({"label": "처음으로", "action": "message", "messageText": "처음으로"})

    return msg, quick_replies, ','.join(path)

# 일반 조사 (직행/바로가기 등 특정 발화에서만 진입)
def find_investigation(places, target):
    df = pd.read_excel(EXCEL_PATH, sheet_name='조사', dtype=str).fillna('')
    cond = (
        (df['장소1'] == places[0]) &
        (df['장소2'] == places[1]) &
        (df['장소3'] == places[2]) &
        (df['장소4'] == places[3]) &
        (df['장소5'] == places[4]) &
        (df['타겟'] == target)
    )
    row = df[cond]
    return row.iloc[0].to_dict() if not row.empty else None

# 정산
def process_settlement(id_code, action, value=None):
    df = pd.read_excel(EXCEL_PATH, sheet_name='정산', dtype=str).fillna('')
    row = df[df['선택2'] == action]
    if not row.empty:
        return row.iloc[0].get('출력', '정산 처리 완료')
    return '정산 처리 완료'

# 랜덤
def get_random_answer(keyword):
    df = pd.read_excel(EXCEL_PATH, sheet_name='랜덤', dtype=str).fillna('')
    row = df[df['랜덤 키워드'] == keyword]
    if not row.empty:
        options = row.iloc[0]['답변 리스트 # , 으로 구별. 이 중 하나를 출력한다.'].split(',')
        return random.choice([opt.strip() for opt in options])
    return '랜덤 응답 없음'

@app.route('/skill', methods=['POST'])
def skill():
    try:
        data = request.json
        params = data.get('action', {}).get('params', {})
        type_ = params.get('type')
        id_code = params.get('id_code', '')

        # 1. 인증
        if type_ == 'auth':
            if not id_code:
                response = "인증코드를 입력해주세요"
            else:
                user = get_user(id_code)
                if not user:
                    response = "인증코드가 유효하지 않습니다."
                else:
                    response = f"{user['Name']}님 어서오세요."
            log_action(id_code, f"[인증] {id_code}", response)
            return jsonify({
                "version": "2.0",
                "template": {
                    "outputs": [{"simpleText": {"text": response}}],
                    "quickReplies": [
                        {"label": "조사", "action": "block", "blockId": "트리형조사블록ID"},
                        {"label": "정산", "action": "block", "blockId": "정산블록ID"}
                    ] if id_code and user else []
                }
            })

        # 2. 트리형 조사 (조사 버튼 클릭 시 조사 시트 첫 행 출력)
        elif type_ == 'investigate_tree':
            select_path = params.get('select_path', '')
            user_input = params.get('user_input', '')
            if not select_path and not user_input:
                # 조사 버튼을 처음 눌렀을 때만
                df = pd.read_excel(EXCEL_PATH, sheet_name='조사', dtype=str).fillna('')
                if df.empty or '출력' not in df.columns:
                    message = "조사 데이터가 없습니다."
                else:
                    message = df.iloc[0]['출력']
                log_action(id_code, "[조사버튼]", message)
                return jsonify({
                    "version": "2.0",
                    "template": {
                        "outputs": [{"simpleText": {"text": message}}]
                    }
                })
            else:
                # 트리형 조사 기존 로직
                msg, quick_replies, new_path = investigate_tree(select_path, user_input)
                log_action(id_code, f"[조사트리] {select_path} + {user_input}", msg)
                return jsonify({
                    "version": "2.0",
                    "template": {
                        "outputs": [{"simpleText": {"text": msg}}],
                        "quickReplies": quick_replies
                    },
                    "context": {
                        "values": [
                            {"name": "select_path", "lifeSpan": 10, "params": {"select_path": new_path}}
                        ]
                    }
                })

        # 3. 일반 조사 (직행/바로가기 등 특정 발화에서만 진입)
        elif type_ == 'investigate':
            user = get_user(id_code)
            utterance = params.get('utterance', '')
            parts = utterance.split()
            if len(parts) < 6 or not user:
                response = "장소와 타겟을 정확히 입력해주세요. 예시: 서울 강남구 삼성동 한국폴리텍 라온관 1층정수기"
                log_action(id_code, utterance, response)
                return jsonify({
                    "version": "2.0",
                    "template": {"outputs": [{"simpleText": {"text": response}}]}
                })
            places, target = parts[:5], parts[5]
            row = find_investigation(places, target)
            if row:
                response = row['출력']
            else:
                response = "해당 조건에 맞는 장소가 없습니다."
            log_action(id_code, utterance, response)
            return jsonify({
                "version": "2.0",
                "template": {"outputs": [{"simpleText": {"text": response}}]}
            })

        # 4. 정산
        elif type_ == 'settle':
            action = params.get('action', '')
            value = params.get('value', None)
            response = process_settlement(id_code, action, value)
            log_action(id_code, f"[정산] {action} {value if value else ''}", response)
            return jsonify({
                "version": "2.0",
                "template": {"outputs": [{"simpleText": {"text": response}}]}
            })

        # 5. 랜덤
        elif type_ == 'random':
            keyword = params.get('keyword', '')
            response = get_random_answer(keyword)
            log_action(id_code, f"[랜덤] {keyword}", response)
            return jsonify({
                "version": "2.0",
                "template": {"outputs": [{"simpleText": {"text": response}}]}
            })

        # 기본
        response = "요청을 이해하지 못했습니다."
        log_action(id_code, str(params), response)
        return jsonify({
            "version": "2.0",
            "template": {"outputs": [{"simpleText": {"text": response}}]}
        })
    except Exception as e:
        error_msg = f"서버 오류: {str(e)}"
        return jsonify({
            "version": "2.0",
            "template": {"outputs": [{"simpleText": {"text": error_msg}}]}
        })

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
